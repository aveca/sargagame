from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
hdr_font = Font(bold=True, color='FFFFFF', size=10)
hdr_fill = PatternFill('solid', fgColor='0D1E1C')
blue = Font(color='0000FF')
center = Alignment(horizontal='center', vertical='center')
border = Border(bottom=Side(style='thin', color='E0E0E0'))
green_bg = PatternFill('solid', fgColor='E8F5E9')
gold_bg = PatternFill('solid', fgColor='FFF8E1')
orange_bg = PatternFill('solid', fgColor='FFF3E0')

# ═══ DASHBOARD ═══
ws = wb.active
ws.title = 'Dashboard'
ws.merge_cells('A1:F1')
ws['A1'] = 'SARGASSES - Business Dashboard'
ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws['A1'].fill = hdr_fill
ws['A1'].alignment = center

kpis = [('MRR (EUR)', "=Revenue!E2"), ('Clients actifs', "=Revenue!D2"), ('Emails', 1), ('Feedbacks', 0), ('Conv. rate', 0.015), ('Sessions/jour', 400)]
for i, (label, val) in enumerate(kpis):
    c = i + 1
    ws.cell(row=3, column=c, value=label).font = Font(bold=True, size=9, color='686868')
    ws.cell(row=3, column=c).alignment = center
    cell = ws.cell(row=4, column=c, value=val)
    cell.font = Font(bold=True, size=20, color='0D1E1C')
    cell.alignment = center
    cell.fill = gold_bg
    if i == 0: cell.number_format = '#,##0.00'
    if i == 4: cell.number_format = '0.0%'

ws['A7'] = 'Objectifs'
ws['A7'].font = Font(bold=True, size=12)
targets = [('', 'Avril', 'Mai', 'Juin', 'Juillet', 'Aout'), ('MRR', 15, 50, 100, 200, 500), ('Clients', 3, 10, 20, 40, 100), ('Emails', 1, 50, 150, 300, 500)]
for r, row in enumerate(targets):
    for c, val in enumerate(row):
        cell = ws.cell(row=8+r, column=c+1, value=val)
        if r == 0: cell.font = Font(bold=True); cell.fill = PatternFill('solid', fgColor='F5F5F5')
        cell.border = border
for c in range(1, 7): ws.column_dimensions[get_column_letter(c)].width = 16

# ═══ REVENUE ═══
ws2 = wb.create_sheet('Revenue')
for i, h in enumerate(['Mois', 'Nouveaux', 'Churned', 'Actifs', 'MRR', 'Couts', 'Net']):
    c = ws2.cell(row=1, column=i+1, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = center

ws2['A2'] = 'Avril 2026'
ws2['B2'] = 3; ws2['B2'].font = blue
ws2['C2'] = 0; ws2['C2'].font = blue
ws2['D2'] = '=B2-C2'
ws2['E2'] = '=D2*4.99'; ws2['E2'].number_format = '#,##0.00'
ws2['F2'] = 1.67; ws2['F2'].font = blue; ws2['F2'].number_format = '#,##0.00'
ws2['G2'] = '=E2-F2'; ws2['G2'].number_format = '#,##0.00'

months = ['Mai', 'Juin', 'Juil', 'Aout', 'Sept', 'Oct', 'Nov', 'Dec', 'Jan 27', 'Fev 27', 'Mars 27']
for m in range(3, 14):
    ws2.cell(row=m, column=1, value=months[m-3] if m-3 < len(months) else '')
    ws2.cell(row=m, column=2).font = blue
    ws2.cell(row=m, column=3).font = blue
    ws2.cell(row=m, column=4, value=f'=D{m-1}+B{m}-C{m}')
    ws2.cell(row=m, column=5, value=f'=D{m}*4.99'); ws2.cell(row=m, column=5).number_format = '#,##0.00'
    ws2.cell(row=m, column=6, value=f'=F{m-1}'); ws2.cell(row=m, column=6).number_format = '#,##0.00'
    ws2.cell(row=m, column=7, value=f'=E{m}-F{m}'); ws2.cell(row=m, column=7).number_format = '#,##0.00'
for c in range(1, 8): ws2.column_dimensions[get_column_letter(c)].width = 14

# ═══ FUNNEL ═══
ws3 = wb.create_sheet('Funnel')
funnel = [('Etape', 'Volume/jour', 'Taux'), ('Sessions', 400, '100%'), ('Onboarding complete', '=B2*0.7', '=B3/B2'), ('Beach view', '=B2*0.45', '=B4/B2'), ('Forecast lock click', '=B4*0.15', '=B5/B2'), ('Premium modal open', '=B5*0.6', '=B6/B2'), ('Stripe click', '=B6*0.3', '=B7/B2'), ('Conversion', '=B7*0.2', '=B8/B2')]
for r, row in enumerate(funnel):
    for c, val in enumerate(row):
        cell = ws3.cell(row=r+1, column=c+1, value=val)
        if r == 0: cell.font = hdr_font; cell.fill = hdr_fill
        if c >= 1 and r > 0: cell.alignment = center
        if c == 2 and r > 0: cell.number_format = '0.00%'
        cell.border = border
for c in range(1, 4): ws3.column_dimensions[get_column_letter(c)].width = 25

# ═══ A/B TESTS ═══
ws4 = wb.create_sheet('A-B Tests')
ab_h = ['Test', 'Var A', 'Var B', 'Sess A', 'Sess B', 'Conv A', 'Conv B', 'Rate A', 'Rate B', 'Lift', 'Winner', 'Status']
for i, h in enumerate(ab_h):
    c = ws4.cell(row=1, column=i+1, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=9); c.fill = hdr_fill

tests = [('lock1', 'control', 'loss'), ('modal1', 'control', 'family'), ('onb1', 'control', 'skip'), ('free1', 'control', 'two_free'), ('vp1', 'feature', 'outcome')]
for r, (tid, va, vb) in enumerate(tests):
    row = r + 2
    ws4.cell(row=row, column=1, value=tid)
    ws4.cell(row=row, column=2, value=va)
    ws4.cell(row=row, column=3, value=vb)
    for c in [4,5,6,7]: ws4.cell(row=row, column=c).font = blue
    ws4.cell(row=row, column=8, value=f'=IF(D{row}=0,"-",F{row}/D{row})'); ws4.cell(row=row, column=8).number_format = '0.00%'
    ws4.cell(row=row, column=9, value=f'=IF(E{row}=0,"-",G{row}/E{row})'); ws4.cell(row=row, column=9).number_format = '0.00%'
    ws4.cell(row=row, column=10, value=f'=IF(OR(D{row}=0,H{row}=0),"-",(I{row}-H{row})/H{row})'); ws4.cell(row=row, column=10).number_format = '0.0%'
    ws4.cell(row=row, column=12, value='Running'); ws4.cell(row=row, column=12).fill = gold_bg
for c in range(1, 13): ws4.column_dimensions[get_column_letter(c)].width = 12

# ═══ ROADMAP ═══
ws5 = wb.create_sheet('Roadmap')
roadmap = [('Cadence', 'Action', 'Status'),
    ('Session', 'Lire daily-metrics.json', 'Auto'), ('Session', 'Checker A/B tests', 'Manuel'), ('Session', 'Appliquer variant gagnant', 'Manuel'),
    ('Quotidien', 'ERDDAP fetch + deploy 4x/jour', 'Auto'), ('Quotidien', 'Stats check + trends', 'Auto'),
    ('Vendredi', 'Email weekend bulletin', 'Auto'), ('Lundi', 'SEO audit + optimize', 'Auto'), ('Jeudi', 'UX report', 'Auto'),
    ('Mensuel', 'Revenue check Stripe', 'Manuel'), ('Mensuel', 'Churn check', 'Manuel'), ('Mensuel', 'Growth emails/clients', 'Manuel'),
    ('Trimestriel', 'PMF survey (40% = PMF)', 'Manuel'), ('Trimestriel', 'Pricing review', 'Manuel'),
    ('Annuel', 'Expansion autres iles', 'Manuel'), ('Annuel', 'Partenariats tourisme', 'Manuel')]
for r, row in enumerate(roadmap):
    for c, val in enumerate(row):
        cell = ws5.cell(row=r+1, column=c+1, value=val)
        if r == 0: cell.font = hdr_font; cell.fill = hdr_fill
        if c == 2 and val == 'Auto': cell.fill = green_bg
        elif c == 2 and val == 'Manuel': cell.fill = orange_bg
        cell.border = border
for c in range(1, 4): ws5.column_dimensions[get_column_letter(c)].width = 35

wb.save('sargasses-tracker.xlsx')
print('OK: sargasses-tracker.xlsx')
